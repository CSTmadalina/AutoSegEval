#last version 
#tubular strcuture evaluation introduced 

import os
import gatetools as gt
import tempfile
import pydicom
import itk
import numpy as np
from math import sqrt
from scipy.spatial.distance import directed_hausdorff
from scipy.spatial import distance
import matplotlib as plt
from scipy.interpolate import Rbf, InterpolatedUnivariateSpline
import matplotlib.pyplot as plt
import xlsxwriter
import os 
from os.path import join as pjoin
from openpyxl import Workbook
import time
import sys
from tqdm import tqdm


rootFolder = "ADMIREauto_contouring/TestSET/"
patientFolders = []
for folder in os.listdir(rootFolder):
    if os.path.isdir(os.path.join(rootFolder,folder)) and folder.startswith("Test"):
        patientFolders.append(folder)
        
#one patient:
patientFolders = ["Test001"]

for folder in patientFolders:
    CTPath = os.path.join(rootFolder, folder, "CT")
    RTSPath = os.path.join(rootFolder, folder, "RTS_Manual_REF")
    RTSPath1 = os.path.join(rootFolder, folder, "RTS_MIM_MV")
    
    #RTSPath1 = "ADMIREauto_contouring/TestSET/Test006/RTS_PATCHFUSION/"
    #RTSPath1 = "ADMIREauto_contouring/TestSET/Test006/RTS_CLB_70DL_CORR/"
    #RTSPath1 = "ADMIREauto_contouring/TestSET/Test015/RTS_CLB_70DL_CORR/"
    #RTSPath1 = "ADMIREauto_contouring/TestSET/Test015/RTS_ARTplan_corr/"
    #RTSPath1 = "ADMIREauto_contouring/TestSET/Test015/RTS_RF_corr/"
    #RTSPath = "ADMIREauto_contouring/TestSET/Test015/RTS_Manual_CTVn/"
    #RTSPath1 = "ADMIREauto_contouring/TestSET/Test002/RTS_RF/"
    #RTSPath = "ADMIREauto_contouring/CTVnDB/CTVnTestSET/Test_018/RTS_Manual_CTVn/"
    #RTSPath1 = "ADMIREauto_contouring/TestSET/Test001/RTS_ADMIRE_DL_corr/"
    #RTSPath1 = "ADMIREauto_contouring/TestSET/Test015/RTS_CLB_70DL/"
    #RTSPath1 = "ADMIREauto_contouring/TestSET/Test006/RTS_CLB_CTVN_24DL/"
    #RTSPath1 = "ADMIREauto_contouring/TestSET/Test015/RTS_MIM_CTVn2-4/"
    #RTSPath = "ADMIREauto_contouring/TCIA_database/Head-Neck_Cetuximab/0522c0057_DICOM/"
    #RTSPath1 = "ADMIREauto_contouring/TCIA_database/Head-Neck_Cetuximab/0522c0057_DICOM/" 
    #CTPath ="ADMIREauto_contouring/TCIA_database/Head-Neck_Cetuximab/0522c0057_DICOM/CT/"


    ctPaths = []
    for root, dirs, files in os.walk(CTPath):
        for file in files:
            if file.endswith(".dcm"):
                ctPaths += [os.path.join(CTPath, file)]  
    ctImage = gt.read_dicom(ctPaths)
    print(len(ctPaths))
    voxel = ctImage.GetSpacing()
    print("voxelspacing=", voxel)
    volumeVoxel = np.prod(ctImage.GetSpacing())


    for root, dirs, files in os.walk(RTSPath):
        for file in files:
            if file.startswith("2.16") and file.endswith(".dcm"):
                structset = pydicom.read_file(os.path.join(root,file), force=True)
                roi_names = gt.list_roinames(structset)           
                patientId = structset[(0x0010, 0x0020)].value    
    print(patientId)


    for root, dirs, files in os.walk(RTSPath1):
        for file in files:
            if file.startswith("RTS_MIM_MV") and file.endswith(".dcm"):
                structset1 = pydicom.read_file(os.path.join(root,file), force=True)
                roi_names1 = gt.list_roinames(structset1)           
                patientId = structset1[(0x0010, 0x0020)].value    
    print(patientId)

    wb = Workbook()
    ws = wb.active
    ws['A2'] = ('PatientID')
    ws['A3'] = (patientId)
    ws['B3'] = ("Structure Name")
    ws['C3'] = ("Volume")
    ws['D3'] = ("Manual RTS")
    ws['E3'] = (" AUTO CORR")
    ws['F3'] = ("Dice")
    ws['G3'] = ("HD")
    ws['H3'] = ("95HD")
    ws['I3'] = ("HD50")
    ws['J3'] = ("VOLsimilaity")

    print( roi_names)
    print( roi_names1)

    print('volume voxel= ',volumeVoxel)
    print(len(roi_names), len(roi_names1))
    sort_roi_names=np.sort(roi_names)
    print(sort_roi_names)

    indexKey = 1
    for i in range(len(sort_roi_names)):
        for j in range(len(roi_names1)): 
            if sort_roi_names[i] == roi_names1[j]:
                aroi = gt.region_of_interest(structset, sort_roi_names[i])
                masks = aroi.get_mask(ctImage, corrected=False)
                outputStats = gt.imageStatistics(input=masks)
                volumeRoi = outputStats["sum"]*volumeVoxel/1000.0
                nbVoxelRoi = gt.imageStatistics(input=masks)["sum"]
                
                aroi1 = gt.region_of_interest(structset1, roi_names1[j])
                masks1 = aroi1.get_mask(ctImage, corrected=False)

                if ("SpinalCord" in sort_roi_names[i]) or ("Esophagus" in sort_roi_names[i]) or ("Trachea" in sort_roi_names[i]):
                    array_mask = itk.array_from_image(masks)
                    array_mask1 = itk.array_from_image(masks1)
                    indexStruct = np.sort(np.where(array_mask != 0)[0])
                    array_mask1[:indexStruct[0],:,:] = 0
                    array_mask1[indexStruct[-1]+1:,:,:] = 0
                    tmpImage = itk.image_from_array(array_mask1)
                    tmpImage.CopyInformation(masks1)
                    masks1 = tmpImage

                itk.imwrite(masks1, os.path.join(RTSPath1, sort_roi_names[i] + ".mhd"))       
                outputStats1 = gt.imageStatistics(input=masks1)
                volumeRoi1 = outputStats1["sum"]*volumeVoxel/1000.0  
                nbVoxelRoi1 = gt.imageStatistics(input=masks1)["sum"]

                masks_float = itk.cast_image_filter(masks, ttype=(itk.Image[itk.UC, 3], itk.Image[itk.F, 3]))
                masks1_float = itk.cast_image_filter(masks1, ttype=(itk.Image[itk.UC, 3], itk.Image[itk.F, 3]))

                #filter = itk.HausdorffDistanceImageFilter[itk.Image[itk.F,3],itk.Image[itk.F,3]].New()
                #filter.SetInput1(masks_float)    
                #filter.SetInput2(masks1_float)
                "filter.SetUseImageSpacing(True)
                "filter.Update() 

                #HD = filter.GetHausdorffDistance()
                #AVHD = filter.GetAverageHausdorffDistance() 

                intersectionArray = np.multiply(itk.array_view_from_image(masks),  itk.array_view_from_image(masks1))
                intersection = np.sum(intersectionArray)
                dice = 2.0*intersection/(nbVoxelRoi + nbVoxelRoi1 )
                distances = {}           
                metrics = [metric.DiceCoefficient(), metric.HausdorffDistance(percentile=100, metric='HDmax'),metric.HausdorffDistance(percentile=95, metric='HD95'),metric.HausdorffDistance(percentile=50, metric='HD50'),metric.VolumeSimilarity(metric='VO')]
                metricHD95 = 0
                metricHD50 = 0
                labels = {1: sort_roi_names[i] }
                evaluator = eval_.SegmentationEvaluator(metrics, labels)
                ground_truth = sitk.ReadImage(os.path.join(RTSPath, sort_roi_names[i] + ".mhd"))
                prediction = sitk.ReadImage(os.path.join(RTSPath1, sort_roi_names[i] + ".mhd"))
                evaluator.evaluate(prediction, ground_truth, "T")
                for r in evaluator.results:
                    if "HDmax" in r.metric:
                        metricHDmax = r.value
                    elif "HD95" in r.metric:
                        metricHD95 = r.value
                for r in evaluator.results:
                    if "HD50" in r.metric:
                        metricHD50 = r.value
                    elif "VO" in r.metric:
                        metricVO = r.value
                    
                #print(metricHDmax, metricHD95, metricHD50, metricVO)
                
               
                
                print(indexKey, sort_roi_names[i], volumeRoi, volumeRoi1, '%.4f' %dice, '%.4f' %metricHDmax, '%.4f' %metricHD95,'%.4f' %metricHD50,'%.4f' %metricVO )
                ws.cell(column = 1, row = 3+indexKey+1, value = indexKey)
                ws.cell(column = 2, row = 3+indexKey+1, value = sort_roi_names[i])
                ws.cell(column = 3, row = 3+indexKey+1, value = 'cm3')
                ws.cell(column = 4, row = 3+indexKey+1, value = '%.4f' %volumeRoi)
                ws.cell(column = 5, row = 3+indexKey+1, value = '%.4f' %volumeRoi1)
                ws.cell(column = 6, row = 3+indexKey+1, value = '%.4f' %dice)
                ws.cell(column = 7, row = 3+indexKey+1, value = '%.4f' %metricHDmax)
                ws.cell(column = 8, row = 3+indexKey+1, value =  '%.4f' %metricHD95)
                ws.cell(column = 9, row = 3+indexKey+1, value = '%.4f' %metricHD50)
                ws.cell(column = 10, row = 3+indexKey+1, value = '%.4f' %metricVO)
                indexKey = indexKey+1



    print(patientId)
#output_path = pjoin( "ADMIREauto_contouring","TestSET" , "Test018", "ManualvsEKTDL" + patientId + ".xlsx" )    
#output_path = pjoin( "ADMIREauto_contouring","TestSET" , "Test019", "AfterCorrection_Artplan_" + patientId + ".xlsx" ) 
#output_path = pjoin( "ADMIREauto_contouring","TestSET" , "Test015", "2AfterCorrection_70DL_" + patientId + ".xlsx" ) 
#output_path = pjoin( "ADMIREauto_contouring","TestSET" , "Test009", "AfterCorrection_admire_Thyroid_" + patientId + ".xlsx" ) 
#output_path = pjoin( "ADMIREauto_contouring","TestSET" , "Test015", "RTS_MIM_CTVN_" + patientId + ".xlsx" )  
#output_path = pjoin( "ADMIREauto_contouring","TCIA_database" , "Head-Neck_Cetuximab", "0522c0057_DICOM", "TCIA0522c0057_"+ patientId + ".xlsx" )  

output_path = pjoin( RTSPath1, "RTS_MIM_MV_" + patientId + ".xlsx" ) 

wb.save(output_path)

print("done")
